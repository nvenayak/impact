import copy
import datetime
import time
import time as sys_time
import pandas as pd
import numpy as np

from openpyxl import load_workbook

from .core.TrialIdentifier import TimeCourseIdentifier
from .core.AnalyteData import Biomass, Substrate, Product, Reporter, TimePoint
from .core import SingleTrial, ReplicateTrial

from warnings import warn


def parse_raw_identifier(raw_identifier, id_type):
    ti = TimeCourseIdentifier()
    if id_type == 'CSV':
        ti.parse_trial_identifier_from_csv(raw_identifier)
    elif id_type == 'traverse':
        ti.parse_identifier(raw_identifier)
    else:
        raise Exception('Unknown identifier type %s' % id_type)

    return ti


def spectromax_OD(experiment, data, id_type='traverse', plate_type='96 Wells'):
    from .core.settings import settings
    live_calculations = settings.live_calculations

    unparsed_identifiers = data['identifiers']
    raw_data = data['data']

    # The data starts at (3,2) and is in a 8x12 format
    timepoint_list = []

    # Parse identifiers (to prevent parsing at every time point)
    identifiers = []
    for i, row in enumerate(unparsed_identifiers):
        parsed_row = []
        for j, data in enumerate(row):
            if unparsed_identifiers[i][j] not in ['', 0, '0', None]:
                temp_trial_identifier = parse_raw_identifier(unparsed_identifiers[i][j], id_type)
                parsed_row.append(temp_trial_identifier)
            else:
                parsed_row.append(None)
        identifiers.append(parsed_row)

    for start_row_index in range(3, len(raw_data), 9):
        if raw_data[start_row_index][0] != '~End':
            if isinstance(raw_data[start_row_index][0], datetime.datetime):
                raise Exception("Imported a datetime object, make sure to set all cells to 'TEXT' if importing"
                                " from excel")
            parsed_time = raw_data[start_row_index][0].split(':')

            # Convert the Spectromax format to raw hours
            if len(parsed_time) > 2:
                time = int(parsed_time[0]) * 3600 + int(parsed_time[1]) * 60 + int(parsed_time[2])
            else:
                time = int(parsed_time[0]) * 60 + int(parsed_time[1])

            time /= 3600  # convert to hours

            # Define the data for a single plate, single timepoint
            plate_data = [row[2:14] for row in raw_data[start_row_index:start_row_index + 8]]

            # Load the data point by point
            for i, row in enumerate(plate_data):
                for j, data in enumerate(row):
                    # Skip wells where no identifier is listed or no data present
                    if identifiers[i][j] is not None and data not in [None, '']:
                        temp_trial_identifier = identifiers[i][j]
                        temp_trial_identifier.analyte_type = 'biomass'
                        temp_trial_identifier.analyte_name = 'OD600'
                        try:
                            temp_timepoint = TimePoint(temp_trial_identifier, time, float(data))
                        except Exception as e:
                            print(plate_data)
                            raise Exception(e)
                        timepoint_list.append(temp_timepoint)
        else:
            break

    replicate_trial_list = parse_time_point_list(timepoint_list)
    for rep in replicate_trial_list:
        experiment.add_replicate_trial(rep)

    if live_calculations:   experiment.calculate()



def HPLC_titer_parser(experiment, data, id_type='CSV', plate_type='96 Wells'):
    t0 = sys_time.time()

    # Parameters
    row_with_titer_names = 0
    row_with_titer_types = 1
    row_with_first_data = 2
    data_sheet_name = "titers"

    # Initialize variables
    analyte_nameColumn = dict()
    titer_type = dict()

    data = data[data_sheet_name]

    for i in range(1, len(data[row_with_titer_names])):
        analyte_nameColumn[data[row_with_titer_names][i]] = i
        titer_type[data[row_with_titer_names][i]] = \
            data[row_with_titer_types][i]

    # Initialize a timepoint_collection for each titer type (column)
    tempTimePointCollection = dict()
    for names in analyte_nameColumn:
        tempTimePointCollection[names] = []
    skipped_lines = 0
    timepoint_list = []
    for i in range(row_with_first_data, len(data)):
        if type(data[i][0]) is str:
            for key in tempTimePointCollection:
                trial_identifier = TimeCourseIdentifier()

                if id_type == 'CSV':
                    trial_identifier.parse_trial_identifier_from_csv(data[i][0])
                elif id_type == 'traverse':
                    trial_identifier.parse_identifier(data[i][0])

                trial_identifier.analyte_name = key
                trial_identifier.analyte_type = titer_type[key]

                if data[i][analyte_nameColumn[key]] == 'nan':
                    data[i][analyte_nameColumn[key]] = np.nan
                timepoint_list.append(
                    TimePoint(trial_identifier=trial_identifier,
                              time=float(trial_identifier.time),
                              data=data[i][analyte_nameColumn[key]]))

        else:
            skipped_lines += 1
    tf = sys_time.time()
    print("Parsed %i timeCourseObjects in %0.3fs" % (len(timepoint_list), tf - t0), end='')
    print("...Number of lines skipped: ", skipped_lines)
    replicate_trial_list = parse_time_point_list(timepoint_list)
    for rep in replicate_trial_list:
        experiment.add_replicate_trial(rep)


def tecan(experiment, data, id_type='traverse', plate_type='96 Wells'):

    from .core.settings import settings
    live_calculations = settings.live_calculations

    # Add new plate layouts here if necessary
    plate_dict = {'96 Wells': {'num_of_wells': 96, 'num_of_columns': 12},
                  '48 Wells': {'num_of_wells': 48, 'num_of_columns': 8},
                  '24 Wells': {'num_of_wells': 24, 'num_of_columns': 6}}

    # Add new analytes/reporters in this dictionary
    mode_dict = {'Absorbance': {600: {'analyte_name': 'OD600', 'analyte_type': 'biomass'},

                                700: {'analyte_name': 'OD700', 'analyte_type': 'biomass'}},

                 'Fluorescence Top Reading': {(488, 525): {'analyte_name': 'GFP',
                                                           'analyte_type': 'reporter'},

                                              (570, 610): {'analyte_name': 'mCherry',
                                                           'analyte_type': 'reporter'}},

                 'Fluorescence Bottom Reading': {(488, 525): {'analyte_name': 'GFP',
                                                              'analyte_type': 'reporter'},

                                                 (570, 610): {'analyte_name': 'mCherry',
                                                              'analyte_type': 'reporter'}}}

    # Initialize variables
    unparsed_identifiers = data['identifiers']
    raw_data = data['data']
    time_row_index = []
    analyte_dict = {}
    num_of_analytes = 0

    for i, row in enumerate(raw_data):
        if row[0] == 'Mode':
            mode = (next((mode for mode in reversed(row) if mode is not None)))
            if mode == 'Absorbance':
                num_of_analytes += 1
                for nextrow in raw_data[i:i + 5]:
                    if nextrow[0] in ['Wavelength', 'Measurement wavelength']:
                        wavelength = (next((wavelength for wavelength in reversed(nextrow) if type(wavelength) == int)))
                        break
                if wavelength in mode_dict[mode].keys():
                    analyte_dict[num_of_analytes] = mode_dict[mode][wavelength]
                else:
                    analyte_dict[num_of_analytes] = {'analyte_name': 'Reporter' + str(wavelength),
                                                     'analyte_type': 'reporter'}
            elif 'Fluorescence' in mode:
                num_of_analytes += 1
                for nextrow in raw_data[i:i + 7]:
                    if nextrow[0] in ['Excitation Wavelength', 'Excitation wavelength']:
                        ex = int(next((wavelength for wavelength in reversed(nextrow) if type(wavelength) in [int,float])))
                    if nextrow[0] in ['Emission wavelength', 'Emission Wavelength']:
                        em = int(next((wavelength for wavelength in reversed(nextrow) if type(wavelength) in [int,float])))
                        break
                if (ex, em) in mode_dict[mode].keys():
                    analyte_dict[num_of_analytes] = mode_dict[mode][(ex, em)]
                else:
                    analyte_dict[num_of_analytes] = {'analyte_name': 'Reporter' + str(ex) + '/' + str(em),
                                                     'analyte_type': 'reporter'}

        elif 'Time [s]' in row:
            time_row_index.append(i)

    timepoint_list = []
    identifiers = []
    for i, row in enumerate(unparsed_identifiers):
        parsed_row = []
        for j, data in enumerate(row):
            if unparsed_identifiers[i][j] not in ['', 0, '0', None]:
                temp_trial_identifier = TimeCourseIdentifier()
                if id_type == 'CSV':
                    temp_trial_identifier.parse_trial_identifier_from_csv(unparsed_identifiers[i][j])
                elif id_type == 'traverse':
                    temp_trial_identifier.parse_identifier(unparsed_identifiers[i][j])
                parsed_row.append(temp_trial_identifier)
            else:
                parsed_row.append(None)
        identifiers.append(parsed_row)
    print('Extracting timepoint data...', end='')
    t0 = sys_time.time()
    for analyte_num in range(num_of_analytes):
        data_start_index = time_row_index[analyte_num] + 2
        num_of_timepoints = len(raw_data[time_row_index[analyte_num]]) - 1
        for i, data_column_index in enumerate(range(1, num_of_timepoints + 1)):
            time = raw_data[time_row_index[analyte_num]][data_column_index]
            time = (time / 3600)
            for j, data_row_index in enumerate(range(data_start_index, data_start_index + plate_dict[plate_type] \
                    ['num_of_wells'])):
                try:
                    identifier = identifiers[int(j / plate_dict[plate_type]['num_of_columns'])] \
                        [int(j % plate_dict[plate_type]['num_of_columns'])]
                except IndexError:
                    identifier_exists = False
                else:
                    identifier_exists = True
                if identifier_exists:
                    if identifier is not None and raw_data[data_row_index][data_column_index] not in [None, '']:
                        temp_trial_identifier = copy.deepcopy(identifiers[int(j / plate_dict[plate_type]['num_of_columns'])] \
                                                                  [int(j % plate_dict[plate_type]['num_of_columns'])])
                        temp_trial_identifier.analyte_type = analyte_dict[analyte_num + 1]['analyte_type']
                        temp_trial_identifier.analyte_name = analyte_dict[analyte_num + 1]['analyte_name']
                        try:
                            temp_timepoint = TimePoint(temp_trial_identifier, time,
                                                       float(raw_data[data_row_index][data_column_index]))
                        except Exception as e:
                            print(raw_data[data_row_index][data_column_index])
                            raise Exception(e)
                        timepoint_list.append(temp_timepoint)

    tf = sys_time.time()
    print("Extracted time point data in %0.1fs" % ((tf - t0)))

    replicate_trial_list = parse_time_point_list(timepoint_list)
    for rep in replicate_trial_list:
        experiment.add_replicate_trial(rep)

    if live_calculations:   experiment.calculate()


class Parser(object):
    parser_case_dict = {}

    @classmethod
    def register_parser(cls, parser_name, parser_method):
        cls.parser_case_dict[parser_name] = parser_method

    @classmethod
    def parse_raw_data(cls, data_format=None, id_type='traverse', file_name=None, data=None, experiment=None, plate_type = '96 Wells'):
        """
        Parses raw data into an experiment object

        Parameters
        ----------
        format (str): spectromax_OD, spectromax_OD_triplicate, default_titers, tecan_OD
        id_type (str): traverse (id1:value|id2:value) or CSV (deprecated)
        file_name (str): path to structured file
        data (str): dictionary containing data with sheets appropriate to parser
        experiment (str): `Experiment` instance to parse data into, will create new instance if None

        Returns
        -------
        `Experiment`
        """
        if experiment is None:
            from .core.Experiment import Experiment
            experiment = Experiment()

        t0 = time.time()
        if data_format is None:
            raise Exception('No format defined')

        if data is None:
            if file_name is None:
                raise Exception('No data or file name given to load data from')

            # Get data from xlsx file
            print('\nImporting data from %s...' % (file_name), end='')
            # data = get_data(file_name)
            xls_data = load_workbook(filename=file_name, data_only=True)
            print('%0.1fs' % (time.time() - t0))

            # Extract data from sheets
            data = {sheet.title: [[elem.value if elem is not None else None for elem in row]
                                  for row in xls_data[sheet.title]] for sheet in xls_data}

        if data_format in cls.parser_case_dict.keys():
            cls.parser_case_dict[data_format](experiment, data=data, id_type=id_type, plate_type = plate_type)
        else:
            raise Exception('Parser %s not found', data_format)

        return experiment


# Register known parsers
Parser.register_parser('spectromax_OD', spectromax_OD)
Parser.register_parser('tecan_OD', tecan)
Parser.register_parser('default_titers', HPLC_titer_parser)
#Parser.register_parser('spectromax_OD_triplicate', spectromax_OD_triplicate)
Parser.register_parser('tecan_OD_GFP_mCherry', tecan)
Parser.register_parser('tecan', tecan)



def parse_raw_data(format=None, id_type='CSV', file_name=None, data=None, experiment=None):
    """
    Parses raw data into an experiment object (deprecated, use class based parser)

    Parameters
    ----------
    format (str): spectromax_OD, spectromax_OD_triplicate, default_titers, tecan_OD
    id_type (str): traverse (id1:value|id2:value) or CSV (deprecated)
    file_name (str): path to structured file
    data (str): dictionary containing data with sheets appropriate to parser
    experiment (str): `Experiment` instance to parse data into, will create new instance if None

    Returns
    -------
    `Experiment`
    """
    warn('Use class based parser Parser.parse_raw_data')

    if experiment is None:
        from .core.Experiment import Experiment
        experiment = Experiment()

    t0 = time.time()
    if format is None:
        raise Exception('No format defined')

    if data is None:
        if file_name is None:
            raise Exception('No data or file name given to load data from')

        # Get data from xlsx file
        print('\nImporting data from %s...' % (file_name), end='')
        # data = get_data(file_name)
        xls_data = load_workbook(filename=file_name, data_only=True)
        print('%0.1fs' % (time.time() - t0))

        # Extract data from sheets
        data = {sheet.title: [[elem.value if elem is not None else None for elem in row]
                              for row in xls_data[sheet.title]] for sheet in xls_data}

    # Import parsers
    parser_case_dict = {'spectromax_OD'           : spectromax_OD,
                        'tecan_OD'                : tecan,
                        'default_titers'          : HPLC_titer_parser,
                        'tecan'                   : tecan
                        #'spectromax_OD_triplicate': spectromax_OD_triplicate
                        }
    if format in parser_case_dict.keys():
        parser_case_dict[format](experiment, data=data, id_type=id_type)
    else:
        raise Exception('Parser %s not found', format)

    return experiment


def parse_analyte_data(analyte_data_list):
    print('Parsing analyte list...', end='')
    t0 = time.time()

    uniques = list(set([titer.trial_identifier.unique_single_trial() for titer in analyte_data_list]))

    single_trial_list = []
    for unique in uniques:
        single_trial = SingleTrial()
        for titer in analyte_data_list:
            if titer.trial_identifier.unique_single_trial() == unique:
                single_trial.add_analyte_data(titer)
        single_trial_list.append(single_trial)

    tf = time.time()
    print("Parsed %i single trials in %0.1fms" % (len(single_trial_list), (tf - t0) * 1000))

    return parse_single_trial_list(single_trial_list)


def parse_time_point_list(time_point_list):
    print('Parsing time point list...', end='')
    t0 = time.time()
    analyte_dict = {}
    for timePoint in time_point_list:
        if timePoint.get_unique_timepoint_id() in analyte_dict:
            analyte_dict[timePoint.get_unique_timepoint_id()].add_timepoint(timePoint)
        else:
            case_dict = {'biomass'  : Biomass,
                         'substrate': Substrate,
                         'product'  : Product,
                         'reporter' : Reporter}
            if timePoint.trial_identifier.analyte_type in case_dict.keys():
                analyte_dict[timePoint.get_unique_timepoint_id()] = \
                    case_dict[timePoint.trial_identifier.analyte_type]()
                analyte_dict[timePoint.get_unique_timepoint_id()].add_timepoint(timePoint)
            else:
                raise Exception('Unexpected analyte type %s' % timePoint.trial_identifier.analyte_type)
    for analyte in analyte_dict.values():
        analyte.pd_series = pd.Series([timePoint.data for timePoint in analyte.time_points],
                                           index=[timePoint.time for timePoint in analyte.time_points])

    tf = time.time()
    print("Parsed %i time points in %0.1fs" % (len(time_point_list), (tf - t0)))
    return parse_analyte_data(list(analyte_dict.values()))


def parse_single_trial_list(single_trial_list):
    print('Parsing single trial list...', end='')
    t0 = time.time()
    uniques = list(set([single_trial.trial_identifier.unique_replicate_trial()
                        for single_trial in single_trial_list]
                       )
                   )
    replicate_trial_list = []
    for unique in uniques:
        related_trials = [single_trial for single_trial in single_trial_list if
                          single_trial.trial_identifier.unique_replicate_trial() == unique]
        replicate_trial = ReplicateTrial()
        for replicate in related_trials:
            replicate_trial.add_replicate(replicate)
        replicate_trial_list.append(replicate_trial)

    tf = time.time()
    print("Parsed %i replicates in %0.1fs" % (len(replicate_trial_list), (tf - t0)))
    return replicate_trial_list
